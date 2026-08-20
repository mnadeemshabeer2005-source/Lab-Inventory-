from flask import Flask, render_template, request, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
from datetime import datetime
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import csv
import io
import os
import uuid


app = Flask(__name__)

db_url = os.environ.get('DATABASE_URL', 'sqlite:///inventory.db')
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql+pg8000://', 1)
if db_url.startswith('postgresql://'):
    db_url = db_url.replace('postgresql://', 'postgresql+pg8000://', 1)


app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'changeme123')
login_manager = LoginManager(app)
login_manager.login_view = 'login'

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='viewer')

    def is_admin(self):
        return self.role == 'admin'

    def is_staff(self):
        return self.role in ['admin', 'staff']

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.context_processor
def inject_counts():
    try:
        item_count = Item.query.count()
        chem_count = Chemical.query.count()
    except:
        item_count = 0
        chem_count = 0
    return dict(item_count=item_count, chem_count=chem_count)

class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    cupboard = db.Column(db.String(50), nullable=False)
    shelf = db.Column(db.String(50))
    box = db.Column(db.String(50))
    expiry_date = db.Column(db.String(20))
    notes = db.Column(db.String(200))

    def __repr__(self):
        return f"<Item {self.name} ({self.cupboard})>"

class Chemical(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    category = db.Column(db.String(100))
    cupboard = db.Column(db.String(50), nullable=False)
    shelf = db.Column(db.String(50))
    box = db.Column(db.String(50))
    quantity = db.Column(db.Integer, nullable=False, default=0)
    volume = db.Column(db.String(50))
    expiry_date = db.Column(db.String(20))
    safety_notes = db.Column(db.String(500))

    def __repr__(self):
        return f"<Chemical {self.name} ({self.cupboard} / {self.shelf} / {self.box})>"

class Document(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.String(500))
    category = db.Column(db.String(100), nullable=False)
    file_url = db.Column(db.String(500), nullable=False)
    public_id = db.Column(db.String(200))
    uploaded_by = db.Column(db.String(80))
    upload_date = db.Column(db.DateTime, default=datetime.now)
    linked_item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=True)
    linked_chemical_id = db.Column(db.Integer, db.ForeignKey('chemical.id'), nullable=True)

    def __repr__(self):
        return f"<Document {self.name}>"

class PurchaseRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    created_by = db.Column(db.String(80))
    created_date = db.Column(db.DateTime, default=datetime.now)
    notes = db.Column(db.String(500))
    items = db.relationship('PurchaseItem', backref='request', lazy=True, cascade='all, delete-orphan')

    def __repr__(self):
        return f"<PurchaseRequest {self.title}>"

class PurchaseItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    request_id = db.Column(db.Integer, db.ForeignKey('purchase_request.id'), nullable=False)
    item_name = db.Column(db.String(300), nullable=False)
    quantity = db.Column(db.String(100))
    order = db.Column(db.Integer, default=0)
    companies = db.relationship('PurchaseCompanyOption', backref='item', lazy=True, cascade='all, delete-orphan')

    def __repr__(self):
        return f"<PurchaseItem {self.item_name}>"

class PurchaseCompanyOption(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('purchase_item.id'), nullable=False)
    company_name = db.Column(db.String(200))
    description = db.Column(db.String(500))
    price = db.Column(db.String(100))
    column_order = db.Column(db.Integer, default=0)


class Resource(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(300))
    color = db.Column(db.String(20), default='#007bff')
    active = db.Column(db.Boolean, default=True)
    bookings = db.relationship('Booking', backref='resource', lazy=True, cascade='all, delete-orphan')

    def __repr__(self):
        return f"<Resource {self.name}>"

class Booking(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    resource_id = db.Column(db.Integer, db.ForeignKey('resource.id'), nullable=False)
    date = db.Column(db.String(20), nullable=False)
    start_time = db.Column(db.String(10), nullable=False)
    end_time = db.Column(db.String(10), nullable=False)
    purpose = db.Column(db.String(500), nullable=False)
    requester_name = db.Column(db.String(100), nullable=False)
    requester_contact = db.Column(db.String(100))
    status = db.Column(db.String(20), default='pending')
    notes = db.Column(db.String(300))
    created_date = db.Column(db.DateTime, default=datetime.now)
    reviewed_by = db.Column(db.String(80))
    reviewed_date = db.Column(db.DateTime)

    def __repr__(self):
        return f"<Booking {self.resource_id} {self.date} {self.start_time}-{self.end_time}>"

class Equipment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    model = db.Column(db.String(100))
    serial_number = db.Column(db.String(100))
    manufacturer = db.Column(db.String(100))
    location = db.Column(db.String(100))
    purchase_date = db.Column(db.String(20))
    warranty_expiry = db.Column(db.String(20))
    status = db.Column(db.String(30), default='Working')
    specs = db.Column(db.String(1000))
    notes = db.Column(db.String(500))
    service_records = db.relationship('ServiceRecord', backref='equipment', lazy=True, cascade='all, delete-orphan')

    def __repr__(self):
        return f"<Equipment {self.name}>"

class ServiceRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'), nullable=False)
    service_date = db.Column(db.String(20), nullable=False)
    description = db.Column(db.String(500), nullable=False)
    performed_by = db.Column(db.String(100))
    next_service_date = db.Column(db.String(20))
    created_by = db.Column(db.String(80))
    created_date = db.Column(db.DateTime, default=datetime.now)

    def __repr__(self):
        return f"<ServiceRecord {self.equipment_id} {self.service_date}>"

class PracticalLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    practical_name = db.Column(db.String(200), nullable=False)
    purpose = db.Column(db.String(500))
    logged_by = db.Column(db.String(80), nullable=False)
    date = db.Column(db.DateTime, default=datetime.now)
    items_used = db.relationship('LogItem', backref='log', lazy=True, cascade='all, delete-orphan')
    chemicals_used = db.relationship('LogChemical', backref='log', lazy=True, cascade='all, delete-orphan')
    equipment_used = db.relationship('LogEquipment', backref='log', lazy=True, cascade='all, delete-orphan')

    def __repr__(self):
        return f"<PracticalLog {self.practical_name}>"

class LogItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    log_id = db.Column(db.Integer, db.ForeignKey('practical_log.id'), nullable=False)
    item_name = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, default=1)

class LogChemical(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    log_id = db.Column(db.Integer, db.ForeignKey('practical_log.id'), nullable=False)
    chemical_name = db.Column(db.String(120), nullable=False)
    quantity = db.Column(db.String(50), nullable=False)

class LogEquipment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    log_id = db.Column(db.Integer, db.ForeignKey('practical_log.id'), nullable=False)
    equipment_name = db.Column(db.String(100), nullable=False)    

    

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('home'))
        return render_template('login.html', error='Invalid username or password')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/setup', methods=['GET', 'POST'])
def setup():
    if User.query.count() > 0:
        return redirect(url_for('home'))
    if request.method == 'POST':
        username = request.form['username']
        password = generate_password_hash(request.form['password'])
        user = User(username=username, password=password, role='admin')
        db.session.add(user)
        db.session.commit()
        return redirect(url_for('login'))
    return render_template('setup.html')

@app.route('/createtables')
def create_tables():
    db.create_all()
    return "Tables created!"

with app.app_context():
    db.create_all()
    try:
        db.session.execute(db.text('ALTER TABLE "user" ADD COLUMN IF NOT EXISTS role VARCHAR(20) DEFAULT \'admin\''))
        db.session.commit()
    except:
        db.session.rollback()

# ---------- MAIN INDEX ----------
@app.route('/')
def index():
    search_term = request.args.get('q', '').strip()
    cupboard_filter = request.args.get('cupboard', '').strip()
    query = Item.query
    if search_term:
        query = query.filter(
            or_(
                Item.name.ilike(f'%{search_term}%'),
                Item.cupboard.ilike(f'%{search_term}%'),
                Item.shelf.ilike(f'%{search_term}%'),
                Item.box.ilike(f'%{search_term}%'),
                Item.notes.ilike(f'%{search_term}%')
            )
        )
    if cupboard_filter:
        query = query.filter(Item.cupboard == cupboard_filter)
    items = query.order_by(Item.name.asc()).all()
    cupboards = [c[0] for c in db.session.query(Item.cupboard).distinct().all() if c[0]]
    return render_template('index.html', items=items, search=search_term,
                           cupboard=cupboard_filter, cupboards=cupboards)

# ---------- CHEMICALS ----------
@app.route('/chemicals')
def chemicals():
    search = request.args.get('search', '').strip()
    cupboard = request.args.get('cupboard', '').strip()
    category = request.args.get('category', '').strip()
    query = Chemical.query
    if search:
        query = query.filter(
            or_(
                Chemical.name.ilike(f"%{search}%"),
                Chemical.cupboard.ilike(f"%{search}%"),
                Chemical.shelf.ilike(f"%{search}%"),
                Chemical.box.ilike(f"%{search}%"),
                Chemical.volume.ilike(f"%{search}%"),
                Chemical.safety_notes.ilike(f"%{search}%"),
                Chemical.category.ilike(f"%{search}%")
            )
        )
    if cupboard:
        query = query.filter(Chemical.cupboard == cupboard)
    if category:
        query = query.filter(Chemical.category == category)
    chems = query.order_by(Chemical.name.asc()).all()
    cupboards = [c[0] for c in db.session.query(Chemical.cupboard).distinct().all()]
    categories = [c[0] for c in db.session.query(Chemical.category).distinct().all() if c[0]]
    return render_template('chemicals.html', chemicals=chems, search=search,
                           cupboard=cupboard, category=category,
                           cupboards=cupboards, categories=categories)

@app.route('/add', methods=['GET', 'POST'])
@login_required
def add_item():
    if not current_user.is_staff():
        return redirect(url_for('home'))
    if request.method == 'POST':
        new_item = Item(
            name=request.form['name'].strip(),
            quantity=int(request.form['quantity']),
            cupboard=request.form['cupboard'].strip(),
            shelf=request.form.get('shelf', '').strip(),
            box=request.form.get('box', '').strip(),
            expiry_date=request.form.get('expiry_date', '').strip(),
            notes=request.form.get('notes', '').strip()
        )
        db.session.add(new_item)
        db.session.commit()
        return redirect(url_for('index'))
    return render_template('add_item.html')

@app.route('/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
def edit_item(item_id):
    if not current_user.is_staff():
        return redirect(url_for('home'))
    item = Item.query.get_or_404(item_id)
    if request.method == 'POST':
        item.name = request.form['name'].strip()
        item.quantity = int(request.form['quantity'])
        item.cupboard = request.form['cupboard'].strip()
        item.shelf = request.form.get('shelf', '').strip()
        item.box = request.form.get('box', '').strip()
        item.expiry_date = request.form.get('expiry_date', '').strip()
        item.notes = request.form.get('notes', '').strip()
        db.session.commit()
        return redirect(url_for('index'))
    return render_template('edit_item.html', item=item)

@app.route('/delete/<int:item_id>')
@login_required
def delete_item(item_id):
    if not current_user.is_admin():
        return redirect(url_for('home'))
    item = Item.query.get_or_404(item_id)
    return render_template('confirm_delete.html',
                           item_name=item.name,
                           confirm_url=url_for('confirm_delete_item', item_id=item_id),
                           cancel_url=url_for('index'))

@app.route('/delete/<int:item_id>/confirm')
@login_required
def confirm_delete_item(item_id):
    if not current_user.is_admin():
        return redirect(url_for('home'))
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return redirect(url_for('index'))

@app.route('/cupboard/<cupboard_name>')
def cupboard_view(cupboard_name):
    items = Item.query.filter_by(cupboard=cupboard_name).order_by(Item.shelf.asc(), Item.name.asc()).all()
    grouped = {}
    for it in items:
        key = it.shelf if (it.shelf and it.shelf.strip()) else "— (no shelf)"
        grouped.setdefault(key, []).append(it)
    def shelf_sort_key(s):
        import re
        m = re.search(r'(\d+)', s)
        return (0, int(m.group(1))) if m else (1, s.lower())
    sorted_shelves = sorted(grouped.keys(), key=shelf_sort_key)
    return render_template('cupboard_view.html', grouped=grouped, shelves=sorted_shelves, cupboard_name=cupboard_name)

@app.route('/export/csv')
@login_required
def export_csv():
    if not current_user.is_staff():
        return redirect(url_for('home'))
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Quantity", "Cupboard", "Shelf", "Box", "Expiry Date", "Notes"])
    for it in Item.query.order_by(Item.name.asc()).all():
        writer.writerow([it.id, it.name, it.quantity, it.cupboard, it.shelf or "", it.box or "", it.expiry_date or "", it.notes or ""])
    mem = io.BytesIO()
    mem.write(output.getvalue().encode('utf-8'))
    mem.seek(0)
    filename = f"lab_inventory_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(mem, as_attachment=True, download_name=filename, mimetype='text/csv')

@app.route('/chemicals/cupboard/<cupboard_name>')
def chemicals_cupboard_view(cupboard_name):
    chems = Chemical.query.filter_by(cupboard=cupboard_name).order_by(Chemical.shelf.asc(), Chemical.name.asc()).all()
    grouped = {}
    for c in chems:
        key = c.shelf if (c.shelf and c.shelf.strip()) else "— (no shelf)"
        grouped.setdefault(key, []).append(c)
    def shelf_sort_key(s):
        import re
        m = re.search(r'(\d+)', s)
        return (0, int(m.group(1))) if m else (1, s.lower())
    sorted_shelves = sorted(grouped.keys(), key=shelf_sort_key)
    return render_template('chemicals_cupboard_view.html', grouped=grouped, shelves=sorted_shelves, cupboard_name=cupboard_name)

@app.route('/chemicals/add', methods=['GET', 'POST'])
@login_required
def add_chemical():
    if not current_user.is_staff():
        return redirect(url_for('home'))
    if request.method == 'POST':
        new = Chemical(
           name=request.form['name'].strip(),
           cupboard=request.form['cupboard'].strip(),
           shelf=request.form.get('shelf', '').strip(),
           box=request.form.get('box', '').strip(),
           quantity=int(request.form.get('quantity', 0)),
           volume=request.form.get('volume', '').strip(),
           expiry_date=request.form.get('expiry_date', '').strip(),
           safety_notes=request.form.get('safety_notes', '').strip(),
           category=request.form.get('category', '').strip()
        )
        db.session.add(new)
        db.session.commit()
        return redirect(url_for('chemicals'))
    return render_template('add_chemical.html')

@app.route('/chemicals/edit/<int:chem_id>', methods=['GET', 'POST'])
@login_required
def edit_chemical(chem_id):
    if not current_user.is_staff():
        return redirect(url_for('home'))
    chem = Chemical.query.get_or_404(chem_id)
    if request.method == 'POST':
        chem.name = request.form['name'].strip()
        chem.category = request.form.get('category', '').strip()
        chem.cupboard = request.form['cupboard'].strip()
        chem.shelf = request.form.get('shelf', '').strip()
        chem.box = request.form.get('box', '').strip()
        chem.quantity = int(request.form.get('quantity', 0))
        chem.volume = request.form.get('volume', '').strip()
        chem.expiry_date = request.form.get('expiry_date', '').strip()
        chem.safety_notes = request.form.get('safety_notes', '').strip()
        db.session.commit()
        return redirect(url_for('chemicals'))
    return render_template('edit_chemical.html', chem=chem)

@app.route('/chemicals/delete/<int:chem_id>')
@login_required
def delete_chemical(chem_id):
    if not current_user.is_admin():
        return redirect(url_for('home'))
    chem = Chemical.query.get_or_404(chem_id)
    return render_template('confirm_delete.html',
                           item_name=chem.name,
                           confirm_url=url_for('confirm_delete_chemical', chem_id=chem_id),
                           cancel_url=url_for('chemicals'))

@app.route('/chemicals/delete/<int:chem_id>/confirm')
@login_required
def confirm_delete_chemical(chem_id):
    if not current_user.is_admin():
        return redirect(url_for('home'))
    chem = Chemical.query.get_or_404(chem_id)
    db.session.delete(chem)
    db.session.commit()
    return redirect(url_for('chemicals'))

@app.route('/export/chemicals/csv')
@login_required
def export_chemicals_csv():
    if not current_user.is_staff():
        return redirect(url_for('home'))
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["ID", "Name", "Cupboard", "Shelf", "Box", "Quantity", "Volume", "Expiry Date", "Safety Notes"])
    for c in Chemical.query.order_by(Chemical.name.asc()).all():
        w.writerow([c.id, c.name, c.cupboard, c.shelf or "", c.box or "", c.quantity, c.volume or "", c.expiry_date or "", c.safety_notes or ""])
    mem = io.BytesIO()
    mem.write(output.getvalue().encode('utf-8'))
    mem.seek(0)
    filename = f"chemicals_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(mem, as_attachment=True, download_name=filename, mimetype='text/csv')

# ---------- HOME ----------
@app.route('/home')
def home():
    total_items = Item.query.count()
    total_chemicals = Chemical.query.count()
    recent_items = Item.query.order_by(Item.id.desc()).limit(5).all()
    recent_chemicals = Chemical.query.order_by(Chemical.id.desc()).limit(5).all()
    expiring_soon = 0
    low_stock = 0
    return render_template('home.html', total_items=total_items, total_chemicals=total_chemicals,
                           expiring_soon=expiring_soon, low_stock=low_stock,
                           recent_items=recent_items, recent_chemicals=recent_chemicals)

# ---------- GLOBAL SEARCH ----------
@app.route('/search')
def global_search():
    search_term = request.args.get('q', '').strip()
    if not search_term:
        return redirect(url_for('home'))
    items = Item.query.filter(
        or_(Item.name.ilike(f'%{search_term}%'), Item.cupboard.ilike(f'%{search_term}%'),
            Item.shelf.ilike(f'%{search_term}%'), Item.box.ilike(f'%{search_term}%'),
            Item.notes.ilike(f'%{search_term}%'))
    ).order_by(Item.name.asc()).all()
    chemicals = Chemical.query.filter(
        or_(Chemical.name.ilike(f'%{search_term}%'), Chemical.cupboard.ilike(f'%{search_term}%'),
            Chemical.shelf.ilike(f'%{search_term}%'), Chemical.box.ilike(f'%{search_term}%'),
            Chemical.volume.ilike(f'%{search_term}%'), Chemical.safety_notes.ilike(f'%{search_term}%'),
            Chemical.category.ilike(f'%{search_term}%'))
    ).order_by(Chemical.name.asc()).all()
    return render_template('search_results.html', search_term=search_term, items=items, chemicals=chemicals)

# ---------- ADMIN ----------
@app.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin():
        return redirect(url_for('home'))
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
def admin_add_user():
    if not current_user.is_admin():
        return redirect(url_for('home'))
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = generate_password_hash(request.form['password'])
        role = request.form.get('role', 'viewer')
        if User.query.filter_by(username=username).first():
            return render_template('admin_add_user.html', error='Username already exists')
        user = User(username=username, password=password, role=role)
        db.session.add(user)
        db.session.commit()
        return redirect(url_for('admin_users'))
    return render_template('admin_add_user.html')

@app.route('/admin/users/delete/<int:user_id>')
@login_required
def admin_delete_user(user_id):
    if not current_user.is_admin():
        return redirect(url_for('home'))
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        return redirect(url_for('admin_users'))
    db.session.delete(user)
    db.session.commit()
    return redirect(url_for('admin_users'))

# ---------- PRINT ----------
@app.route('/print/items')
@login_required
def print_items():
    if not current_user.is_staff():
        return redirect(url_for('home'))
    search_term = request.args.get('q', '').strip()
    cupboard_filter = request.args.get('cupboard', '').strip()
    query = Item.query
    if search_term:
        query = query.filter(
            or_(
                Item.name.ilike(f'%{search_term}%'),
                Item.cupboard.ilike(f'%{search_term}%'),
                Item.shelf.ilike(f'%{search_term}%'),
                Item.box.ilike(f'%{search_term}%'),
                Item.notes.ilike(f'%{search_term}%')
            )
        )
    if cupboard_filter:
        query = query.filter(Item.cupboard == cupboard_filter)
    items = query.order_by(Item.cupboard.asc(), Item.name.asc()).all()
    now = datetime.now().strftime('%d %B %Y, %I:%M %p')
    return render_template('print_items.html', items=items, now=now,
                           search=search_term, cupboard=cupboard_filter)
@app.route('/print/chemicals')
@login_required
def print_chemicals():
    if not current_user.is_staff():
        return redirect(url_for('home'))
    search = request.args.get('search', '').strip()
    cupboard = request.args.get('cupboard', '').strip()
    category = request.args.get('category', '').strip()
    query = Chemical.query
    if search:
        query = query.filter(
            or_(
                Chemical.name.ilike(f"%{search}%"),
                Chemical.cupboard.ilike(f"%{search}%"),
                Chemical.shelf.ilike(f"%{search}%"),
                Chemical.box.ilike(f"%{search}%"),
                Chemical.volume.ilike(f"%{search}%"),
                Chemical.safety_notes.ilike(f"%{search}%"),
                Chemical.category.ilike(f"%{search}%")
            )
        )
    if cupboard:
        query = query.filter(Chemical.cupboard == cupboard)
    if category:
        query = query.filter(Chemical.category == category)
    chemicals = query.order_by(Chemical.cupboard.asc(), Chemical.name.asc()).all()
    now = datetime.now().strftime('%d %B %Y, %I:%M %p')
    return render_template('print_chemicals.html', chemicals=chemicals, now=now,
                           search=search, cupboard=cupboard, category=category)

# ---------- DOCUMENTS ----------
DOCUMENT_CATEGORIES = ['Haematology', 'Biochemistry', 'Procedures', 'Safety', 'Equipment', 'Other']

@app.route('/documents')
def documents():
    category = request.args.get('category', '').strip()
    search = request.args.get('search', '').strip()
    query = Document.query
    if category:
        query = query.filter(Document.category == category)
    if search:
        query = query.filter(
            or_(Document.name.ilike(f'%{search}%'), Document.description.ilike(f'%{search}%'))
        )
    docs = query.order_by(Document.upload_date.desc()).all()
    # Get categories dynamically from database
    categories = [c[0] for c in db.session.query(Document.category).distinct().all() if c[0]]
    return render_template('documents.html', documents=docs, categories=categories,
                           selected_category=category, search=search)

# ---------- Google Drive helper ----------
def get_drive_service():
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    SCOPES = ['https://www.googleapis.com/auth/drive']
    SERVICE_ACCOUNT_FILE = os.path.join(os.path.dirname(__file__), 'lab-inventory-497101-9c4256d8ed51.json')
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build('drive', 'v3', credentials=credentials)

DRIVE_FOLDER_ID = '1oth9qAFWS2iTM4IjbPkNFDjhnVGdsM4S'

@app.route('/documents/upload', methods=['GET', 'POST'])
@login_required
def upload_document():
    if not current_user.is_staff():
        return redirect(url_for('documents'))
    if request.method == 'POST':
        file = request.files.get('file')
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        category = request.form.get('category', '').strip()
        linked_item_id = request.form.get('linked_item_id') or None
        linked_chemical_id = request.form.get('linked_chemical_id') or None

        if file and name and category:
            upload_folder = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
            os.makedirs(upload_folder, exist_ok=True)
            ext = os.path.splitext(file.filename)[1]
            filename = str(uuid.uuid4()) + ext
            filepath = os.path.join(upload_folder, filename)
            file.save(filepath)
            doc = Document(
                name=name,
                description=description,
                category=category,
                file_url=url_for('static', filename=f'uploads/{filename}', _external=True),
                public_id=filename,
                uploaded_by=current_user.username,
                linked_item_id=linked_item_id,
                linked_chemical_id=linked_chemical_id
            )
            db.session.add(doc)
            db.session.commit()
            return redirect(url_for('documents'))

    items = Item.query.order_by(Item.name.asc()).all()
    chemicals = Chemical.query.order_by(Chemical.name.asc()).all()
    return render_template('upload_document.html', categories=DOCUMENT_CATEGORIES,
                           items=items, chemicals=chemicals)


@app.route('/documents/delete/<int:doc_id>')
@login_required
def delete_document(doc_id):
    if not current_user.is_admin():
        return redirect(url_for('documents'))
    doc = Document.query.get_or_404(doc_id)
    try:
        filepath = os.path.join(os.path.dirname(__file__), 'static', 'uploads', doc.public_id)
        if os.path.exists(filepath):
            os.remove(filepath)
    except:
        pass
    db.session.delete(doc)
    db.session.commit()
    return redirect(url_for('documents'))









# ---------- PURCHASE REQUESTS ----------
@app.route('/purchases')
@login_required
def purchases():
    reqs = PurchaseRequest.query.order_by(PurchaseRequest.created_date.desc()).all()
    return render_template('purchases.html', requests=reqs)

@app.route('/purchases/new', methods=['GET', 'POST'])
@login_required
def new_purchase():
    if not current_user.is_staff():
        return redirect(url_for('purchases'))
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        notes = request.form.get('notes', '').strip()
        pr = PurchaseRequest(title=title, notes=notes, created_by=current_user.username)
        db.session.add(pr)
        db.session.flush()

        item_names = request.form.getlist('item_name')
        item_quantities = request.form.getlist('item_quantity')

        for i, item_name in enumerate(item_names):
            if not item_name.strip():
                continue
            pi = PurchaseItem(
                request_id=pr.id,
                item_name=item_name.strip(),
                quantity=item_quantities[i] if i < len(item_quantities) else '',
                order=i
            )
            db.session.add(pi)
            db.session.flush()

            company_names = request.form.getlist(f'company_name_{i}')
            descriptions = request.form.getlist(f'description_{i}')
            prices = request.form.getlist(f'price_{i}')

            for j, company in enumerate(company_names):
                if not company.strip():
                    continue
                opt = PurchaseCompanyOption(
                    item_id=pi.id,
                    company_name=company.strip(),
                    description=descriptions[j] if j < len(descriptions) else '',
                    price=prices[j] if j < len(prices) else '',
                    column_order=j
                )
                db.session.add(opt)

        db.session.commit()
        return redirect(url_for('view_purchase', req_id=pr.id))
    return render_template('new_purchase.html')


@app.route('/purchases/<int:req_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_purchase(req_id):
    if not current_user.is_staff():
        return redirect(url_for('purchases'))
    pr = PurchaseRequest.query.get_or_404(req_id)
    if request.method == 'POST':
        pr.title = request.form.get('title', '').strip()
        pr.notes = request.form.get('notes', '').strip()

        for item in pr.items:
            for company in item.companies:
                db.session.delete(company)
            db.session.delete(item)
        db.session.flush()

        item_names = request.form.getlist('item_name')
        item_quantities = request.form.getlist('item_quantity')

        for i, item_name in enumerate(item_names):
            if not item_name.strip():
                continue
            pi = PurchaseItem(
                request_id=pr.id,
                item_name=item_name.strip(),
                quantity=item_quantities[i] if i < len(item_quantities) else '',
                order=i
            )
            db.session.add(pi)
            db.session.flush()

            company_names = request.form.getlist(f'company_name_{i}')
            descriptions = request.form.getlist(f'description_{i}')
            prices = request.form.getlist(f'price_{i}')

            for j, company in enumerate(company_names):
                if not company.strip():
                    continue
                opt = PurchaseCompanyOption(
                    item_id=pi.id,
                    company_name=company.strip(),
                    description=descriptions[j] if j < len(descriptions) else '',
                    price=prices[j] if j < len(prices) else '',
                    column_order=j
                )
                db.session.add(opt)

        db.session.commit()
        return redirect(url_for('view_purchase', req_id=pr.id))
    return render_template('edit_purchase.html', pr=pr)


@app.route('/purchases/<int:req_id>/export/excel')
@login_required
def export_purchase_excel(req_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    pr = PurchaseRequest.query.get_or_404(req_id)
    max_companies = max((len(item.companies) for item in pr.items), default=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Request"

    header_font = Font(bold=True, color="000000", size=11)
    header_fill = PatternFill("solid", fgColor="FFD700")
    item_fill = PatternFill("solid", fgColor="FFFDE7")
    qty_fill = PatternFill("solid", fgColor="EBF5FB")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Title row
    total_cols = max_companies + 2
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    title_cell = ws['A1']
    title_cell.value = pr.title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center

    # Date row
    ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
    date_cell = ws['A2']
    date_cell.value = f"Created by {pr.created_by} · {pr.created_date.strftime('%d %B %Y')}"
    date_cell.font = Font(italic=True, size=10)
    date_cell.alignment = center

    # Header row
    headers = ['ITEMS TO BE PURCHASED', 'QUANTITY'] + [f'COMPANY {i+1}' for i in range(max_companies)]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # Data rows
    for row_idx, item in enumerate(pr.items, start=4):
        item_cell = ws.cell(row=row_idx, column=1, value=f"{row_idx - 3}. {item.item_name}")
        item_cell.font = Font(bold=True, size=10)
        item_cell.fill = item_fill
        item_cell.border = border
        item_cell.alignment = left

        qty_cell = ws.cell(row=row_idx, column=2, value=item.quantity or '-')
        qty_cell.fill = qty_fill
        qty_cell.border = border
        qty_cell.alignment = center
        qty_cell.font = Font(size=10)

        for col_idx in range(max_companies):
            cell = ws.cell(row=row_idx, column=col_idx + 3)
            if col_idx < len(item.companies):
                company = item.companies[col_idx]
                parts = []
                if company.company_name:
                    parts.append(company.company_name)
                if company.description:
                    parts.append(company.description)
                if company.price:
                    parts.append(company.price)
                cell.value = ' - '.join(parts)
            cell.border = border
            cell.alignment = left
            cell.font = Font(size=10)

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    for i in range(max_companies):
        ws.column_dimensions[get_column_letter(i + 3)].width = 40

    # Row heights
    for row in ws.iter_rows(min_row=4, max_row=3 + len(pr.items)):
        for cell in row:
            ws.row_dimensions[cell.row].height = 50

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    filename = f"purchase_{pr.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(mem, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/purchases/<int:req_id>')
@login_required
def view_purchase(req_id):
    pr = PurchaseRequest.query.get_or_404(req_id)
    max_companies = 0
    for item in pr.items:
        if len(item.companies) > max_companies:
            max_companies = len(item.companies)
    return render_template('view_purchase.html', pr=pr, max_companies=max_companies)

@app.route('/purchases/<int:req_id>/print')
@login_required
def print_purchase(req_id):
    pr = PurchaseRequest.query.get_or_404(req_id)
    max_companies = 0
    for item in pr.items:
        if len(item.companies) > max_companies:
            max_companies = len(item.companies)
    now = datetime.now().strftime('%d %B %Y')
    return render_template('print_purchase.html', pr=pr, max_companies=max_companies, now=now)

@app.route('/purchases/<int:req_id>/delete')
@login_required
def delete_purchase(req_id):
    if not current_user.is_admin():
        return redirect(url_for('purchases'))
    pr = PurchaseRequest.query.get_or_404(req_id)
    db.session.delete(pr)
    db.session.commit()
    return redirect(url_for('purchases'))

# ---------- CALENDAR ----------
@app.route('/calendar')
def calendar_view():
    resources = Resource.query.filter_by(active=True).all()
    bookings = Booking.query.filter_by(status='approved').order_by(Booking.date.asc(), Booking.start_time.asc()).all()
    bookings_data = []
    for b in bookings:
        bookings_data.append({
            'id': b.id,
            'title': f"{b.resource.name} - {b.requester_name}",
            'date': b.date,
            'start_time': b.start_time,
            'end_time': b.end_time,
            'purpose': b.purpose,
            'color': b.resource.color,
            'resource': b.resource.name
        })
    import json
    return render_template('calendar.html', resources=resources,
                           bookings_json=json.dumps(bookings_data))

@app.route('/calendar/request', methods=['GET', 'POST'])
def booking_request():
    resources = Resource.query.filter_by(active=True).all()
    if request.method == 'POST':
        resource_id = request.form.get('resource_id')
        date = request.form.get('date')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        purpose = request.form.get('purpose', '').strip()
        requester_name = request.form.get('requester_name', '').strip()
        requester_contact = request.form.get('requester_contact', '').strip()

        conflicts = Booking.query.filter_by(
            resource_id=resource_id,
            date=date,
            status='approved'
        ).all()

        has_conflict = False
        for c in conflicts:
            if not (end_time <= c.start_time or start_time >= c.end_time):
                has_conflict = True
                break

        if has_conflict:
            from datetime import date as date_type
            now_date = date_type.today().isoformat()
            return render_template('booking_request.html', resources=resources,
                                   error='This time slot conflicts with an existing booking. Please choose another time.',
                                   prefill=request.form, now_date=now_date)

        booking = Booking(
            resource_id=resource_id,
            date=date,
            start_time=start_time,
            end_time=end_time,
            purpose=purpose,
            requester_name=requester_name,
            requester_contact=requester_contact,
            status='pending'
        )
        db.session.add(booking)
        db.session.commit()
        return redirect(url_for('booking_submitted'))

    prefill = request.args
    from datetime import date as date_type
    now_date = date_type.today().isoformat()
    return render_template('booking_request.html', resources=resources,
                           prefill=prefill, now_date=now_date)

@app.route('/calendar/submitted')
def booking_submitted():
    return render_template('booking_submitted.html')

@app.route('/calendar/admin')
@login_required
def booking_admin():
    if not current_user.is_staff():
        return redirect(url_for('calendar_view'))
    pending = Booking.query.filter_by(status='pending').order_by(Booking.date.asc()).all()
    approved = Booking.query.filter_by(status='approved').order_by(Booking.date.desc()).limit(20).all()
    rejected = Booking.query.filter_by(status='rejected').order_by(Booking.date.desc()).limit(10).all()
    resources = Resource.query.all()
    return render_template('booking_admin.html', pending=pending, approved=approved,
                           rejected=rejected, resources=resources)

@app.route('/calendar/booking/<int:booking_id>/approve')
@login_required
def approve_booking(booking_id):
    if not current_user.is_staff():
        return redirect(url_for('calendar_view'))
    booking = Booking.query.get_or_404(booking_id)
    booking.status = 'approved'
    booking.reviewed_by = current_user.username
    booking.reviewed_date = datetime.now()
    db.session.commit()
    return redirect(url_for('booking_admin'))

@app.route('/calendar/booking/<int:booking_id>/reject')
@login_required
def reject_booking(booking_id):
    if not current_user.is_staff():
        return redirect(url_for('calendar_view'))
    booking = Booking.query.get_or_404(booking_id)
    booking.status = 'rejected'
    booking.reviewed_by = current_user.username
    booking.reviewed_date = datetime.now()
    db.session.commit()
    return redirect(url_for('booking_admin'))

@app.route('/calendar/booking/<int:booking_id>/delete')
@login_required
def delete_booking(booking_id):
    if not current_user.is_admin():
        return redirect(url_for('booking_admin'))
    booking = Booking.query.get_or_404(booking_id)
    db.session.delete(booking)
    db.session.commit()
    return redirect(url_for('booking_admin'))

# ---------- RESOURCES ----------
@app.route('/calendar/resources')
@login_required
def manage_resources():
    if not current_user.is_admin():
        return redirect(url_for('calendar_view'))
    resources = Resource.query.all()
    return render_template('manage_resources.html', resources=resources)

@app.route('/calendar/resources/add', methods=['GET', 'POST'])
@login_required
def add_resource():
    if not current_user.is_admin():
        return redirect(url_for('calendar_view'))
    if request.method == 'POST':
        resource = Resource(
            name=request.form.get('name', '').strip(),
            description=request.form.get('description', '').strip(),
            color=request.form.get('color', '#007bff')
        )
        db.session.add(resource)
        db.session.commit()
        return redirect(url_for('manage_resources'))
    return render_template('add_resource.html')

@app.route('/calendar/resources/delete/<int:resource_id>')
@login_required
def delete_resource(resource_id):
    if not current_user.is_admin():
        return redirect(url_for('manage_resources'))
    resource = Resource.query.get_or_404(resource_id)
    db.session.delete(resource)
    db.session.commit()
    return redirect(url_for('manage_resources'))
# ---------- EQUIPMENT ----------
@app.route('/equipment')
@login_required
def equipment():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    query = Equipment.query
    if search:
        query = query.filter(
            or_(
                Equipment.name.ilike(f'%{search}%'),
                Equipment.model.ilike(f'%{search}%'),
                Equipment.serial_number.ilike(f'%{search}%'),
                Equipment.manufacturer.ilike(f'%{search}%'),
                Equipment.location.ilike(f'%{search}%')
            )
        )
    if status:
        query = query.filter(Equipment.status == status)
    equipments = query.order_by(Equipment.name.asc()).all()
    return render_template('equipment.html', equipments=equipments,
                           search=search, status=status)

@app.route('/equipment/add', methods=['GET', 'POST'])
@login_required
def add_equipment():
    if not current_user.is_staff():
        return redirect(url_for('equipment'))
    if request.method == 'POST':
        eq = Equipment(
            name=request.form.get('name', '').strip(),
            model=request.form.get('model', '').strip(),
            serial_number=request.form.get('serial_number', '').strip(),
            manufacturer=request.form.get('manufacturer', '').strip(),
            location=request.form.get('location', '').strip(),
            purchase_date=request.form.get('purchase_date', '').strip(),
            warranty_expiry=request.form.get('warranty_expiry', '').strip(),
            status=request.form.get('status', 'Working').strip(),
            specs=request.form.get('specs', '').strip(),
            notes=request.form.get('notes', '').strip()
        )
        db.session.add(eq)
        db.session.commit()
        return redirect(url_for('equipment'))
    return render_template('add_equipment.html')

@app.route('/equipment/<int:eq_id>')
@login_required
def view_equipment(eq_id):
    eq = Equipment.query.get_or_404(eq_id)
    return render_template('view_equipment.html', eq=eq)

@app.route('/equipment/<int:eq_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_equipment(eq_id):
    if not current_user.is_staff():
        return redirect(url_for('equipment'))
    eq = Equipment.query.get_or_404(eq_id)
    if request.method == 'POST':
        eq.name = request.form.get('name', '').strip()
        eq.model = request.form.get('model', '').strip()
        eq.serial_number = request.form.get('serial_number', '').strip()
        eq.manufacturer = request.form.get('manufacturer', '').strip()
        eq.location = request.form.get('location', '').strip()
        eq.purchase_date = request.form.get('purchase_date', '').strip()
        eq.warranty_expiry = request.form.get('warranty_expiry', '').strip()
        eq.status = request.form.get('status', 'Working').strip()
        eq.specs = request.form.get('specs', '').strip()
        eq.notes = request.form.get('notes', '').strip()
        db.session.commit()
        return redirect(url_for('view_equipment', eq_id=eq.id))
    return render_template('edit_equipment.html', eq=eq)

@app.route('/equipment/<int:eq_id>/delete')
@login_required
def delete_equipment(eq_id):
    if not current_user.is_admin():
        return redirect(url_for('equipment'))
    eq = Equipment.query.get_or_404(eq_id)
    db.session.delete(eq)
    db.session.commit()
    return redirect(url_for('equipment'))

@app.route('/equipment/<int:eq_id>/service/add', methods=['GET', 'POST'])
@login_required
def add_service_record(eq_id):
    if not current_user.is_staff():
        return redirect(url_for('view_equipment', eq_id=eq_id))
    eq = Equipment.query.get_or_404(eq_id)
    if request.method == 'POST':
        record = ServiceRecord(
            equipment_id=eq_id,
            service_date=request.form.get('service_date', '').strip(),
            description=request.form.get('description', '').strip(),
            performed_by=request.form.get('performed_by', '').strip(),
            next_service_date=request.form.get('next_service_date', '').strip(),
            created_by=current_user.username
        )
        db.session.add(record)
        db.session.commit()
        return redirect(url_for('view_equipment', eq_id=eq_id))
    return render_template('add_service_record.html', eq=eq)

@app.route('/equipment/service/<int:record_id>/delete')
@login_required
def delete_service_record(record_id):
    if not current_user.is_admin():
        return redirect(url_for('equipment'))
    record = ServiceRecord.query.get_or_404(record_id)
    eq_id = record.equipment_id
    db.session.delete(record)
    db.session.commit()
    return redirect(url_for('view_equipment', eq_id=eq_id))
@app.route('/equipment/export/excel')
@login_required
def export_equipment_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    equipments = Equipment.query.order_by(Equipment.name.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment List"

    header_font = Font(bold=True, color="000000", size=11)
    header_fill = PatternFill("solid", fgColor="007BFF")
    white_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Lab Equipment List — {datetime.now().strftime('%d %B %Y')}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center

    # Headers
    headers = ['Name', 'Model', 'Serial Number', 'Manufacturer', 'Location', 'Status', 'Purchase Date', 'Warranty Expiry']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # Status fill colors
    status_fills = {
        'Working': PatternFill("solid", fgColor="D4EDDA"),
        'Under Maintenance': PatternFill("solid", fgColor="FFF3CD"),
        'Broken': PatternFill("solid", fgColor="F8D7DA"),
        'Decommissioned': PatternFill("solid", fgColor="E2E3E5"),
    }

    # Data rows
    for row_idx, eq in enumerate(equipments, start=3):
        values = [eq.name, eq.model or '-', eq.serial_number or '-',
                  eq.manufacturer or '-', eq.location or '-', eq.status,
                  eq.purchase_date or '-', eq.warranty_expiry or '-']
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = left
            cell.font = Font(size=10)
            if col_idx == 6:  # Status column
                cell.fill = status_fills.get(eq.status, PatternFill())

    # Column widths
    widths = [30, 20, 20, 20, 20, 20, 15, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    filename = f"equipment_list_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(mem, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Practical Logs ────────────────────────────────────────────────────────────
@app.route('/logs')
def logs():
    all_logs = PracticalLog.query.order_by(PracticalLog.date.desc()).all()
    return render_template('logs.html', logs=all_logs)
@app.route('/logs/add', methods=['GET', 'POST'])
def add_log():
    items = Item.query.order_by(Item.name).all()
    chemicals = Chemical.query.order_by(Chemical.name).all()
    equipment = Equipment.query.order_by(Equipment.name).all()
    if request.method == 'POST':
        practical_name = request.form.get('practical_name', '').strip()
        purpose = request.form.get('purpose', '').strip()
        logged_by = request.form.get('logged_by', '').strip() or 'Guest'
        if not practical_name:
            return render_template('add_log.html', items=items, chemicals=chemicals, equipment=equipment, error='Practical name is required.')
        log = PracticalLog(
            practical_name=practical_name,
            purpose=purpose,
            logged_by=logged_by
        )
        db.session.add(log)
        db.session.flush()
        # Items
        item_names = request.form.getlist('item_name[]')
        item_qtys = request.form.getlist('item_qty[]')
        for name, qty in zip(item_names, item_qtys):
            if name.strip():
                db.session.add(LogItem(log_id=log.id, item_name=name.strip(), quantity=int(qty or 1)))
        # Chemicals
        chem_names = request.form.getlist('chem_name[]')
        chem_qtys = request.form.getlist('chem_qty[]')
        for name, qty in zip(chem_names, chem_qtys):
            if name.strip():
                db.session.add(LogChemical(log_id=log.id, chemical_name=name.strip(), quantity=qty.strip() or '1'))
        # Equipment
        equip_names = request.form.getlist('equip_name[]')
        for name in equip_names:
            if name.strip():
                db.session.add(LogEquipment(log_id=log.id, equipment_name=name.strip()))
        db.session.commit()
        return redirect(url_for('logs'))
    return render_template('add_log.html', items=items, chemicals=chemicals, equipment=equipment)

@app.route('/logs/delete/<int:log_id>')
def delete_log(log_id):
    log = PracticalLog.query.get_or_404(log_id)
    db.session.delete(log)
    db.session.commit()
    return redirect(url_for('logs'))



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
